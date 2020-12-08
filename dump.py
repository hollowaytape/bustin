"""
    Dumps TTB text into an excel sheet for translation and reinsertion.
"""

import sys
import os
import openpyxl
import xlsxwriter
from rominfo import FILES, FILE_BLOCKS, inline_CTRL, DUMP_XLS_PATH, BACKUP_XLS_PATH, FILE_CATEGORIES
from mcv import compress

ASCII_MODE = 1
# 0 = none
# 1: punctuation and c format strings only (not implemented)
# 2: All ascii

THRESHOLD = 4

#FILES = ["TBS.EXE", "decompressed\\SEN013R1.MCV",]
#print(FILES)

existing_strings = {}

def safe_sheet_name(filename):
    sheet_name = filename.split("/")[-1]
    sheet_name = sheet_name.split("\\")[-1]
    sheet_name = sheet_name.replace("decompressed\\", "")
    return sheet_name


def dump(files):
    # Create the overview worksheet now so it shows up first.
    # Can't create it fully right away because links wouldn't work.
    overview_worksheet = workbook.add_worksheet("Overview")

    for filename in files:
        print(filename)
        file_path = os.path.join('original/', filename)

        sheet_name = safe_sheet_name(filename)

        row = 1
        worksheet = workbook.add_worksheet(sheet_name)
        worksheet.write(0, 0, 'Offset', header)
        worksheet.write(0, 1, 'Compressed Offset', header)
        worksheet.write(0, 2, 'Japanese', header)
        worksheet.write(0, 3, 'JP_Len', header)
        worksheet.write(0, 4, 'English', header)
        worksheet.write(0, 5, 'EN_Len', header)
        worksheet.write(0, 6, 'Comments', header)

        worksheet.set_column('A:A', 8)
        worksheet.set_column('B:B', 8)
        worksheet.set_column('C:C', 60)
        worksheet.set_column('D:D', 5)
        worksheet.set_column('E:E', 60)
        worksheet.set_column('F:F', 5)
        worksheet.set_column('G:G', 80)


        if "RTMS" in filename or "MSGS" in filename or "DAT" in filename:
            ASCII_MODE = 0
        else:
            ASCII_MODE = 2

        with open(file_path, 'rb') as f:
            contents = f.read()

        if filename.endswith(".MCV"):
            with open(filename.replace("decompressed\\", "original\\SCN\\"), 'rb') as g:
                compressed_cursor = 0
                compressed_contents = g.read()

        if filename not in FILE_BLOCKS:
            blocks = [(0, len(contents))]
        else:
            blocks = FILE_BLOCKS[filename]

        for block in blocks:
            block_contents = contents[block[0]:block[1]]
            cursor = 0
            sjis_buffer = b""
            sjis_buffer_start = 0
            sjis_strings = []

            while cursor < len(block_contents):

                # First byte of SJIS text. Read the next one, too
                if (0x80 <= block_contents[cursor] <= 0x9f or 0xe0 <= block_contents[cursor] <= 0xef) and cursor+1 < len(block_contents):
                    #print(bytes(block_contents[cursor]))
                    sjis_char = block_contents[cursor:cursor+2]
                    try:
                        ch = sjis_char.decode('shift-jis')
                        sjis_buffer += sjis_char
                    except UnicodeDecodeError:
                        print("Invalid char")
                        sjis_strings.append((sjis_buffer_start+block[0], sjis_buffer))
                        sjis_buffer = b""
                        sjis_buffer_start = cursor+1
                    cursor += 1

                    #print(sjis_char)
                    #sjis_buffer += block_contents[cursor].to_bytes(1, byteorder='little')
                    #cursor += 1
                    #sjis_buffer += block_contents[cursor].to_bytes(1, byteorder='little')

                # ASCII text
                elif 0x20 <=block_contents[cursor] <= 0x7e and ASCII_MODE in (1, 2):
                    sjis_buffer += block_contents[cursor].to_bytes(1, byteorder='little')

                elif block_contents[cursor] in inline_CTRL and len(sjis_buffer) > 0:
                    sjis_buffer += inline_CTRL[block_contents[cursor]]

                # End of continuous SJIS string, so add the buffer to the strings and reset buffer
                else:
                    sjis_strings.append((sjis_buffer_start+block[0], sjis_buffer))
                    sjis_buffer = b""
                    sjis_buffer_start = cursor+1
                cursor += 1
                #print(sjis_buffer)

                if any([sjis_buffer.endswith(x) for x in (b'[END]', b'[WAIT]', b'[LN]', b'[CLR]')]):
                    sjis_buffer = sjis_buffer.replace(b'[END]', b'')
                    sjis_buffer = sjis_buffer.replace(b'[WAIT]', b'')
                    sjis_buffer = sjis_buffer.replace(b'[LN]', b'')
                    sjis_buffer = sjis_buffer.replace(b'[CLR]', b'')
                    for ct in inline_CTRL:
                        if sjis_buffer.endswith(inline_CTRL[ct]):
                            sjis_buffer = sjis_buffer.replace(inline_CTRL[ct], b'')
                    sjis_strings.append((sjis_buffer_start+block[0], sjis_buffer))
                    sjis_buffer = b''
                    sjis_buffer_start = cursor

            # Catch anything left after exiting the loop
            if sjis_buffer:
                sjis_strings.append((sjis_buffer_start+block[0], sjis_buffer))


            if len(sjis_strings) == 0:
                continue

            last_comp_loc = None
            for s in sjis_strings:
                if len(s[1]) < THRESHOLD:
                    continue

                loc = '0x' + hex(s[0]).lstrip('0x').zfill(5)
                try:
                    jp = s[1].decode('shift-jis')
                except UnicodeDecodeError:
                    problem_string = s[1]
                    ind = 0
                    while problem_string:
                        char = problem_string[:2]
                        print(char)
                        print(ind, char.decode('shift-jis'))
                        ind += 2
                        problem_string = problem_string[2:]
                    print(s[1])
                    jp = "Can't decode"
                    print("Couldn't decode that")
                    #continue

                if len(jp.strip()) == 0:
                    continue

                if '=' in jp:
                    jp = jp.replace('=', '[=]')

                print(loc, jp)

                if filename.endswith(".MCV"):
                    comp = compress(s[1])
                    substr = compressed_contents[compressed_cursor:]
                    i = substr.find(comp) + compressed_cursor
                    if last_comp_loc is not None:
                        assert i >= last_comp_loc, "%s is at %s, below cursor %s" % (jp, hex(i), hex(last_comp_loc))
                    last_comp_loc = i
                    compressed_cursor = i + len(comp)
                    compressed_loc = '0x' + hex(i).lstrip('0x').zfill(5)
                    worksheet.write(row, 1, compressed_loc)

                worksheet.write(row, 0, loc)
                worksheet.write(row, 2, jp)
                worksheet.write(row, 3, '=LEN(C%s)' % str(row+1))
                worksheet.write(row, 5, '=LEN(E%s)' % str(row+1))
                #worksheet.write(row, 3, filename)

                if (filename, loc) in existing_strings:
                    print("String already exists")
                    worksheet.write(row, 4, existing_strings[(filename, loc)])

                row += 1

    # Fill in overview worksheet
    worksheet = overview_worksheet
    worksheet.set_column('A:A', 30)
    worksheet.set_column('B:B', 8)
    worksheet.set_column('C:C', 8)
    row = 0
    worksheet.write(row, 0, "Tokyo Twilight Busters", header)
    worksheet.write(row, 1, '=SUM(B3:B10000)')
    worksheet.write(row, 2, '=SUM(C3:C10000)')
    row += 1

    for category in FILE_CATEGORIES:
        row += 1
        worksheet.write(row, 0, category, header)
        worksheet.write(row, 1, "Translated", header)
        worksheet.write(row, 2, "Total", header)
        row += 1

        for filename in FILE_CATEGORIES[category]:
            worksheet.write(row, 0, filename)
            worksheet.write(row, 1, '=COUNTIF(%s!E2:E10000, "*")' % filename)
            worksheet.write(row, 2, '=COUNTIF(%s!C2:C10000, "*")' % filename)
            row += 1

    # Add leftover filenames to an "Unknown" section
    row += 1
    worksheet.write(row, 0, "Uncategorized", header)
    worksheet.write(row, 1, "Translated", header)
    worksheet.write(row, 2, "Total", header)
    row += 1

    for filename in files:
        sheet_name = safe_sheet_name(filename)

        found = False
        for category in FILE_CATEGORIES:
            if sheet_name in FILE_CATEGORIES[category]:
                found = True
        if not found:
            worksheet.write(row, 0, sheet_name)
            worksheet.write(row, 1, '=COUNTIF(%s!E2:E10000, "*")' % sheet_name)
            worksheet.write(row, 2, '=COUNTIF(%s!C2:C10000, "*")' % sheet_name)
            row += 1


    workbook.close()

if __name__ == '__main__':
    for workbook_path in [DUMP_XLS_PATH, BACKUP_XLS_PATH]:
        existing_workbook = openpyxl.load_workbook(BACKUP_XLS_PATH)
        for sheet_name in existing_workbook.sheetnames:
            #print(sheet_name)
            sheet = existing_workbook[sheet_name]
            for row in list(sheet.rows)[1:]:
                #for i, cell in enumerate(row):
                #    print(row[i].value)
                offset = row[0].value
                english = row[4].value
                if english:
                    existing_strings[(sheet_name, offset)] = english

    print(existing_strings)


    workbook = xlsxwriter.Workbook(DUMP_XLS_PATH)
    header = workbook.add_format({'bold': True, 'align': 'center', 'bottom': True, 'bg_color': 'gray'})
    dump(FILES)
